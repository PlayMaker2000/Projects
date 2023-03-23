import pyodbc 
import openpyxl
import pandas as pd
import warnings
warnings.filterwarnings('ignore')

# Connecting to SQL Database
conn = pyodbc.connect('Driver={SQL Server};'
                      'Server=DESKTOP-S1K7KAK\SQLEXPRESS;'
                      'Database=players;'
                      'Trusted_Connection=yes;')
cursor = conn.cursor()

# Opening the stat sheet
workbook = openpyxl.load_workbook("NFL Statsheet.xlsx", data_only=True)
sheet = workbook.active

# Starting point on spreadsheet for team 1 QB
newRow = 5
newColumn = 2

# Find team 1 QB's stats
# List to store scanned data
team1qblist = []
newRow = 5
newColumn = 2
playernamecell = sheet.cell(newRow,newColumn)
while newColumn < 10:
    # If no value is in the "player name" cell
    if playernamecell.value == None:
        print("Player not found")
        break
    cell = sheet.cell(newRow,newColumn)
    team1qblist.append(cell.value)
    # Stop at specified column
    if newColumn == 7:
        break
    newColumn += 1     

# Make scanned data from spreadsheet in team1qblist into Python variables.
team1qbname = team1qblist[0]
team1qbcompletions = team1qblist[1]
team1qbattempts = team1qblist[2]
team1qbpyards = team1qblist[3]
team1qbtouchdowns = team1qblist[4]
team1qbinterceptions = team1qblist[5]

# Methods for updating SQL table
def updatepyards():
    cursor.execute("select yards from Quarterbacks WHERE playername = (?)", (team1qbname))
    realqbyards = 0
    for qbyardrows in cursor:
        # Convert SQL query to a Python variable
        sqlqbyards = str(qbyardrows)
        sqlqbyards = sqlqbyards.replace(",","").replace(" ","").replace("(","").replace(")","")
        sqlqbyards = int(sqlqbyards)
        # Add SQL data and Spreadsheet data together
        realqbyards += sqlqbyards + team1qbpyards
        # Update SQL table
        print(realqbyards)
        cursor.execute("Update Quarterbacks SET yards = (?) WHERE playername = (?)", (realqbyards, team1qbname))
        conn.commit()
        break

def updatetouchdowns():
    cursor.execute("select touchdowns from Quarterbacks WHERE playername = (?)", (team1qbname))
    realqbtouchdowns = 0
    for qbtdrows in cursor:
        # Convert SQL query to a Python variable
        sqlqbtouchdowns = str(qbtdrows)
        sqlqbtouchdowns = sqlqbtouchdowns.replace(",","").replace(" ","").replace("(","").replace(")","")
        sqlqbtouchdowns = int(sqlqbtouchdowns)
        # Add SQL data and Spreadsheet data together
        realqbtouchdowns += sqlqbtouchdowns + team1qbtouchdowns
        # Update SQL table
        print(realqbtouchdowns)
        cursor.execute("Update Quarterbacks SET touchdowns = (?) WHERE playername = (?)", (realqbtouchdowns, team1qbname))
        conn.commit()
        break

def updateinterceptions():
    cursor.execute("select interceptions from Quarterbacks WHERE playername = (?)", (team1qbname))
    realqbints = 0
    for qbintrows in cursor:
        # Convert SQL query to a Python variable
        sqlqbints = str(qbintrows)
        sqlqbints = sqlqbints.replace(",","").replace(" ","").replace("(","").replace(")","")
        sqlqbints = int(sqlqbints)
        # Add SQL data and Spreadsheet data together
        realqbints += sqlqbints + team1qbinterceptions
        # Update SQL table
        print(realqbints)
        cursor.execute("Update Quarterbacks SET interceptions = (?) WHERE playername = (?)", (realqbints, team1qbname))
        conn.commit()
        break

# Check to see if any records exist in table
cursor.execute("IF EXISTS (select * from Quarterbacks) SELECT 'Table not empty' ELSE SELECT 'Table is empty'")
exists = ''
for row in cursor:
    exists = str(row)
    # If table is empty insert values
    if exists == "('Table is empty', )":
        print("It is empty")
        cursor.execute("INSERT INTO Quarterbacks VALUES(?, ?, ?, ?)", (team1qbname, team1qbpyards, team1qbtouchdowns, team1qbinterceptions))
        conn.commit()
        break
    else:
        break


# Check for duplicate names
cursor.execute("select playername from Quarterbacks")
# Make python string into SQL string format to find matching name
sqlplayername = "('{qb1name}', )".format(qb1name = team1qbname)
playername = ''
for row in cursor:
    # Check to see if data existed in table to avoid updating inserted data
    if exists == "('Table not empty', )":
        playername = str(row)
        if playername == sqlplayername:
            # Update Data in Quarterback table if name was found
            print("Matches with name")
            realqbyards = 0
            realqbtouchdowns = 0
            realqbints = 0
            updatepyards()
            updatetouchdowns()
            updateinterceptions()
            break
        # Check to see if data did not exist before to avoid updating inserted data
    if exists == "('Table is empty', )":
        break
    if playername != sqlplayername:
        # Insert data into Quarterback table in SQL if no same names were found
        print("No matching names")
        cursor.execute("INSERT INTO Quarterbacks VALUES(?, ?, ?, ?)", (team1qbname, team1qbpyards, team1qbtouchdowns, team1qbinterceptions))
        conn.commit()
        break
        #cursor.execute("select playername from Quarterbacks")

# Output table from SQL for confirmation
RawData = pd.read_sql_query('''select * from Quarterbacks''', conn)
print(RawData)


